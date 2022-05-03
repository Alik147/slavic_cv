from pathlib import Path
from openpyxl import load_workbook
from books.models import Book

def run(path: str):
	wb = load_workbook(Path(path))
	sheet = wb.active
	flag = False
	for row in sheet.iter_rows():
		cells = row
		if flag == False:
			flag = True
			continue
		if row[0].value is None:
			break
		new_book = Book(storage=row[0].value,
						fond=row[1].value,
						inventory=row[2].value,
						number=row[3].value,
						exodus=row[4].value,
						months=None if row[5].value is None else row[5].value.split(','),
						months_numbers=None if row[6].value is None else row[6].value.split(','),
						special_name=row[7].value,
						dating=row[8].value,
						size=row[9].value,
						book_format=row[10].value,
						book_type=row[11].value,
						file_path=row[12].value if row[12].hyperlink is None else row[12].hyperlink.target,
						handwriting=row[13].value,
						material=row[14].value,
						description=row[15].value if row[12].hyperlink is None else row[12].hyperlink.target,
						additional_info=row[16].value
						)
		new_book.save()
